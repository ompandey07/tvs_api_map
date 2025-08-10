from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.authtoken.models import Token
from rest_framework.pagination import PageNumberPagination
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.db.models import Q
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from .models import BlogModel, MailModel, Customer
from .serializers import (
    LoginSerializer, UserSerializer, BlogSerializer, BlogListSerializer,
    MailSerializer, MailListSerializer, MailToggleSerializer,
    CustomerSerializer, CustomerListSerializer, CustomerImportSerializer,
    DuplicateHandlingSerializer, DashboardStatsSerializer, SearchSerializer,
    FileOperationSerializer
)


# ======================================================================
# CUSTOM PAGINATION CLASSES
# ======================================================================

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 15
    page_size_query_param = 'page_size'
    max_page_size = 100


class SmallResultsSetPagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = 'page_size'
    max_page_size = 50


# ======================================================================
# AUTHENTICATION API VIEWS
# ======================================================================

class LoginAPIView(APIView):
    """
    API view for user authentication
    """
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """
        Authenticate user and return token
        """
        # Default admin credentials (same logic as original)
        DEFAULT_EMAIL = 'admin@tvs.com'
        DEFAULT_PASSWORD = 'admin@1200'
        DEFAULT_USERNAME = 'admin'
        DEFAULT_FIRST_NAME = 'Admin'
        DEFAULT_LAST_NAME = 'User'
        
        # Ensure default admin user exists
        try:
            with transaction.atomic():
                admin_user, created = User.objects.get_or_create(
                    username=DEFAULT_USERNAME,
                    defaults={
                        'email': DEFAULT_EMAIL,
                        'first_name': DEFAULT_FIRST_NAME,
                        'last_name': DEFAULT_LAST_NAME,
                        'is_staff': True,
                        'is_superuser': True,
                        'is_active': True,
                    }
                )
                
                if created or not admin_user.check_password(DEFAULT_PASSWORD):
                    admin_user.set_password(DEFAULT_PASSWORD)
                    admin_user.email = DEFAULT_EMAIL
                    admin_user.save()
                    
        except Exception as e:
            return Response({
                'success': False,
                'error': 'System error occurred. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Validate login data
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']
            
            # Create or get token
            token, created = Token.objects.get_or_create(user=user)
            
            # Login user (for session-based compatibility)
            login(request, user)
            
            return Response({
                'success': True,
                'message': f'Welcome back, {user.first_name or user.username}!',
                'token': token.key,
                'user': UserSerializer(user).data
            }, status=status.HTTP_200_OK)
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


class LogoutAPIView(APIView):
    """
    API view for user logout
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        """
        Logout user and delete token
        """
        try:
            # Delete the user's token
            if hasattr(request.user, 'auth_token'):
                request.user.auth_token.delete()
            
            # Logout user from session
            logout(request)
            
            return Response({
                'success': True,
                'message': 'Successfully logged out'
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ======================================================================
# DASHBOARD API VIEW
# ======================================================================

class DashboardAPIView(APIView):
    """
    API view for admin dashboard statistics
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        """
        Get dashboard statistics
        """
        try:
            selected_template = MailModel.objects.filter(selected_for=True).first()
            
            stats = {
                'total_customers': Customer.objects.count(),
                'sms_sent_count': Customer.objects.filter(sms_status=True).count(),
                'sms_pending_count': Customer.objects.filter(sms_status=False).count(),
                'total_blogs': BlogModel.objects.count(),
                'total_sms_templates': MailModel.objects.count(),
                'selected_sms_template': selected_template.subject if selected_template else None,
            }
            
            serializer = DashboardStatsSerializer(stats)
            
            return Response({
                'success': True,
                'data': serializer.data,
                'user': UserSerializer(request.user).data
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ======================================================================
# BLOG API VIEWSET
# ======================================================================

class BlogViewSet(viewsets.ModelViewSet):
    """
    ViewSet for blog management with search functionality
    """
    queryset = BlogModel.objects.all().order_by('-created_at')
    serializer_class = BlogSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    pagination_class = StandardResultsSetPagination
    
    def get_serializer_class(self):
        """Use different serializers for list and detail views"""
        if self.action == 'list':
            return BlogListSerializer
        return BlogSerializer
    
    def get_queryset(self):
        """Filter blogs based on search query"""
        queryset = BlogModel.objects.all().order_by('-created_at')
        search_query = self.request.query_params.get('search', '')
        
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) | 
                Q(content__icontains=search_query)
            )
        
        return queryset
    
    def destroy(self, request, *args, **kwargs):
        """Delete blog with image file cleanup"""
        try:
            blog = self.get_object()
            blog_title = blog.title
            
            # Delete image file if exists
            if blog.image and os.path.isfile(blog.image.path):
                os.remove(blog.image.path)
            
            blog.delete()
            
            return Response({
                'success': True,
                'message': f'Blog "{blog_title}" has been deleted successfully.'
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error deleting blog: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def update(self, request, *args, **kwargs):
        """Update blog with image file management"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            old_image = instance.image
            
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            if serializer.is_valid():
                # If new image is uploaded, delete old one
                if 'image' in request.FILES and old_image and os.path.isfile(old_image.path):
                    os.remove(old_image.path)
                
                serializer.save()
                
                return Response({
                    'success': True,
                    'message': f'Blog "{serializer.instance.title}" has been updated successfully.',
                    'data': serializer.data
                }, status=status.HTTP_200_OK)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error updating blog: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def create(self, request, *args, **kwargs):
        """Create new blog"""
        try:
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                
                return Response({
                    'success': True,
                    'message': f'Blog "{serializer.instance.title}" has been published successfully.',
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error creating blog: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)


# ======================================================================
# SMS/MAIL API VIEWSET
# ======================================================================

class MailViewSet(viewsets.ModelViewSet):
    """
    ViewSet for SMS/Mail template management
    """
    queryset = MailModel.objects.all().order_by('-created_at')
    serializer_class = MailSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    pagination_class = SmallResultsSetPagination
    
    def get_serializer_class(self):
        """Use different serializers for list and detail views"""
        if self.action == 'list':
            return MailListSerializer
        return MailSerializer
    
    def get_queryset(self):
        """Filter mails based on search query"""
        queryset = MailModel.objects.all().order_by('-created_at')
        search_query = self.request.query_params.get('search', '')
        
        if search_query:
            queryset = queryset.filter(
                Q(subject__icontains=search_query) | 
                Q(content__icontains=search_query) |
                Q(signature__icontains=search_query)
            )
        
        return queryset
    
    def destroy(self, request, *args, **kwargs):
        """Delete mail with template file cleanup"""
        try:
            mail = self.get_object()
            mail_subject = mail.subject
            
            # Delete template file if exists
            if mail.template and os.path.isfile(mail.template.path):
                os.remove(mail.template.path)
            
            mail.delete()
            
            return Response({
                'success': True,
                'message': f'SMS "{mail_subject}" has been deleted successfully.'
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error deleting SMS: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def update(self, request, *args, **kwargs):
        """Update mail with template file management"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            old_template = instance.template
            
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            if serializer.is_valid():
                # If new template is uploaded, delete old one
                if 'template' in request.FILES and old_template and os.path.isfile(old_template.path):
                    os.remove(old_template.path)
                
                serializer.save()
                
                return Response({
                    'success': True,
                    'message': f'SMS "{serializer.instance.subject}" has been updated successfully.',
                    'data': serializer.data
                }, status=status.HTTP_200_OK)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error updating SMS: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def create(self, request, *args, **kwargs):
        """Create new mail"""
        try:
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                
                return Response({
                    'success': True,
                    'message': f'SMS "{serializer.instance.subject}" has been saved successfully.',
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error creating SMS: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def toggle_selected(self, request, pk=None):
        """Toggle selected status for SMS template"""
        try:
            mail = self.get_object()
            serializer = MailToggleSerializer(data=request.data)
            
            if serializer.is_valid():
                selected = serializer.validated_data['selected']
                
                if selected:
                    # Check if another SMS is already selected
                    currently_selected = MailModel.objects.filter(selected_for=True).first()
                    if currently_selected and currently_selected.id != mail.id:
                        return Response({
                            'success': False,
                            'already_selected': True,
                            'current_sms': {
                                'id': currently_selected.id,
                                'subject': currently_selected.subject
                            }
                        }, status=status.HTTP_409_CONFLICT)
                    
                    # Select this SMS and unselect all others
                    MailModel.objects.all().update(selected_for=False)
                    mail.selected_for = True
                    mail.save()
                    
                    return Response({
                        'success': True,
                        'action': 'selected',
                        'message': f'SMS "{mail.subject}" has been selected.'
                    }, status=status.HTTP_200_OK)
                else:
                    # Unselect this SMS
                    mail.selected_for = False
                    mail.save()
                    
                    return Response({
                        'success': True,
                        'action': 'unselected',
                        'message': f'SMS "{mail.subject}" has been unselected.'
                    }, status=status.HTTP_200_OK)
            
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def confirm_switch(self, request, pk=None):
        """Force switch to new SMS template"""
        try:
            mail = self.get_object()
            
            # Unselect all SMS
            MailModel.objects.all().update(selected_for=False)
            
            # Select the new SMS
            mail.selected_for = True
            mail.save()
            
            return Response({
                'success': True,
                'switched': True,
                'message': f'Successfully switched to SMS "{mail.subject}".'
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ======================================================================
# CUSTOMER SMS MANAGEMENT API VIEW
# ======================================================================

class CustomerSMSAPIView(APIView):
    """
    API view for customer SMS management with import/export functionality
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    pagination_class = StandardResultsSetPagination
    
    def get(self, request):
        """
        Get customers list or handle export operations
        """
        action = request.query_params.get('action')
        
        if action == 'export_format':
            return self.export_format_template()
        elif action == 'export_data':
            return self.export_customers_data()
        
        # Default: Return customers list
        return self.get_customers_list(request)
    
    def post(self, request):
        """
        Handle import operations
        """
        action = request.data.get('action')
        
        if action == 'import_data':
            return self.import_customers_data(request)
        elif action == 'handle_duplicates':
            return self.handle_duplicates(request)
        
        return Response({
            'success': False,
            'error': f'Invalid action: {action}'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    def get_customers_list(self, request):
        """Get paginated customers list with search"""
        try:
            search_query = request.query_params.get('search', '')
            customers = Customer.objects.all().order_by('-created_at')
            
            if search_query:
                customers = customers.filter(
                    Q(full_name__icontains=search_query) |
                    Q(contact_number__icontains=search_query) |
                    Q(email__icontains=search_query) |
                    Q(address__icontains=search_query)
                )
            
            # Pagination
            paginator = StandardResultsSetPagination()
            paginated_customers = paginator.paginate_queryset(customers, request)
            serializer = CustomerListSerializer(paginated_customers, many=True)
            
            return paginator.get_paginated_response({
                'success': True,
                'customers': serializer.data,
                'stats': {
                    'total_customers': Customer.objects.count(),
                    'sms_sent_count': Customer.objects.filter(sms_status=True).count(),
                    'sms_pending_count': Customer.objects.filter(sms_status=False).count(),
                },
                'search_query': search_query,
            })
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def export_format_template(self):
        """Export Excel format template with green headers"""
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Customer Data Template"
            
            # Define headers
            headers = ['FULL NAME', 'CONTACT NUMBER', 'ADDRESS', 'EMAIL']
            
            # Add headers to first row
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                
                # Style the header cells
                cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add sample data
            sample_data = [
                ['John Doe', '1234567890', '123 Main St, City, State', 'john.doe@example.com'],
                ['Jane Smith', '0987654321', '456 Oak Ave, Town, State', 'jane.smith@example.com'],
                ['Mike Johnson', '1122334455', '789 Pine Rd, Village, State', 'mike.johnson@example.com'],
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = Font(color="666666")
            
            # Adjust column widths
            column_widths = [25, 20, 35, 30]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Customer_Import_Template.xlsx"'
            
            wb.save(response)
            return response
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error creating template: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def export_customers_data(self):
        """Export current customers data to Excel"""
        try:
            customers = Customer.objects.all().order_by('-created_at')
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Customers Data"
            
            # Define headers
            headers = ['FULL NAME', 'CONTACT NUMBER', 'ADDRESS', 'EMAIL', 'SMS STATUS', 'CREATED AT', 'UPDATED AT', 'SENT SMS AT']
            
            # Add headers to first row
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add customer data
            for row_idx, customer in enumerate(customers, 2):
                ws.cell(row=row_idx, column=1, value=customer.full_name)
                ws.cell(row=row_idx, column=2, value=customer.contact_number)
                ws.cell(row=row_idx, column=3, value=customer.address)
                ws.cell(row=row_idx, column=4, value=customer.email)
                ws.cell(row=row_idx, column=5, value="Sent" if customer.sms_status else "Pending")
                ws.cell(row=row_idx, column=6, value=customer.created_at.strftime('%Y-%m-%d %H:%M:%S') if customer.created_at else '')
                ws.cell(row=row_idx, column=7, value=customer.updated_at.strftime('%Y-%m-%d %H:%M:%S') if customer.updated_at else '')
                ws.cell(row=row_idx, column=8, value=customer.sent_sms_at.strftime('%Y-%m-%d %H:%M:%S') if customer.sent_sms_at else '')
            
            # Adjust column widths
            column_widths = [25, 20, 35, 30, 15, 20, 20, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
            
            # Create response
            timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Customers_Data_{timestamp}.xlsx"'
            
            wb.save(response)
            return response
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error exporting data: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def import_customers_data(self, request):
        """Import customers data from Excel"""
        try:
            serializer = CustomerImportSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    'success': False,
                    'errors': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            excel_file = serializer.validated_data['excel_file']
            
            # Read Excel file
            df = pd.read_excel(excel_file)
            
            # Validate required columns
            required_columns = ['FULL NAME', 'CONTACT NUMBER', 'ADDRESS', 'EMAIL']
            df.columns = df.columns.str.upper().str.strip()
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response({
                    'success': False,
                    'error': f'Missing required columns: {", ".join(missing_columns)}. Found columns: {", ".join(df.columns)}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Remove empty rows
            df = df.dropna(subset=required_columns, how='all')
            
            if df.empty:
                return Response({
                    'success': False,
                    'error': 'No valid data found in the Excel file'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check for duplicates
            duplicates = []
            valid_rows = []
            
            for index, row in df.iterrows():
                try:
                    # Clean data
                    full_name = str(row['FULL NAME']).strip()
                    contact_number = str(row['CONTACT NUMBER']).strip()
                    address = str(row['ADDRESS']).strip()
                    email = str(row['EMAIL']).strip()
                    
                    # Validate required fields
                    if not all([full_name, contact_number, address, email]) or any(val == 'nan' for val in [full_name, contact_number, address, email]):
                        continue
                    
                    # Check for duplicate contact number
                    existing_customer = Customer.objects.filter(contact_number=contact_number).first()
                    
                    row_data = {
                        'row_number': index + 2,
                        'full_name': full_name,
                        'contact_number': contact_number,
                        'address': address,
                        'email': email
                    }
                    
                    if existing_customer:
                        duplicates.append({
                            **row_data,
                            'existing_name': existing_customer.full_name,
                            'existing_email': existing_customer.email,
                            'existing_id': existing_customer.id
                        })
                    else:
                        valid_rows.append(row_data)
                        
                except Exception as e:
                    continue
            
            # Store data in session for later processing
            request.session['import_data'] = {
                'valid_rows': valid_rows,
                'duplicates': duplicates,
                'total_rows': len(df),
                'imported_by': request.user.username
            }
            
            if duplicates:
                return Response({
                    'success': True,
                    'has_duplicates': True,
                    'duplicates': duplicates,
                    'valid_count': len(valid_rows),
                    'duplicate_count': len(duplicates),
                    'total_rows': len(df)
                }, status=status.HTTP_200_OK)
            else:
                # No duplicates, proceed with import
                return self.process_import(request, valid_rows, [])
                
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error reading Excel file: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def handle_duplicates(self, request):
        """Handle duplicate resolution and process import"""
        try:
            serializer = DuplicateHandlingSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    'success': False,
                    'errors': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            action = serializer.validated_data['duplicate_action']
            
            if action == 'cancel':
                if 'import_data' in request.session:
                    del request.session['import_data']
                return Response({
                    'success': True,
                    'action': 'cancelled'
                }, status=status.HTTP_200_OK)
            
            import_data = request.session.get('import_data')
            if not import_data:
                return Response({
                    'success': False,
                    'error': 'No import data found'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            valid_rows = import_data['valid_rows']
            duplicates = import_data['duplicates']
            
            # Process based on action
            if action == 'replace':
                return self.process_import(request, valid_rows, duplicates, replace=True)
            elif action == 'ignore':
                return self.process_import(request, valid_rows, [])
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def process_import(self, request, valid_rows, duplicates, replace=False):
        """Process the actual import with progress tracking"""
        try:
            total_operations = len(valid_rows) + (len(duplicates) if replace else 0)
            processed = 0
            created_count = 0
            updated_count = 0
            errors = []
            
            # Process valid rows (new customers)
            for row_data in valid_rows:
                try:
                    Customer.objects.create(
                        full_name=row_data['full_name'],
                        contact_number=row_data['contact_number'],
                        address=row_data['address'],
                        email=row_data['email']
                    )
                    created_count += 1
                except Exception as e:
                    errors.append(f"Row {row_data['row_number']}: {str(e)}")
                
                processed += 1
            
            # Process duplicates if replacing
            if replace and duplicates:
                for dup_data in duplicates:
                    try:
                        Customer.objects.filter(id=dup_data['existing_id']).update(
                            full_name=dup_data['full_name'],
                            address=dup_data['address'],
                            email=dup_data['email'],
                        )
                        updated_count += 1
                    except Exception as e:
                        errors.append(f"Row {dup_data['row_number']}: {str(e)}")
                    
                    processed += 1
            
            # Clean up session
            if 'import_data' in request.session:
                del request.session['import_data']
            
            return Response({
                'success': True,
                'created_count': created_count,
                'updated_count': updated_count,
                'total_processed': processed,
                'errors': errors,
                'message': f'Successfully processed {processed} records. Created: {created_count}, Updated: {updated_count}',
                'imported_by': request.user.username
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error during import: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ======================================================================
# DOCUMENTATION API VIEW
# ======================================================================

class DocumentationAPIView(APIView):
    """
    API view for documentation
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        """
        Get API documentation information
        """
        documentation = {
            'title': 'Admin Dashboard API Documentation',
            'version': '1.0.0',
            'description': 'Comprehensive API for managing blogs, SMS templates, and customers',
            'endpoints': {
                'authentication': {
                    'login': '/api/login/',
                    'logout': '/api/logout/',
                },
                'dashboard': '/api/dashboard/',
                'blogs': '/api/blogs/',
                'sms_templates': '/api/mails/',
                'customers': '/api/customers/',
                'documentation': '/api/documentation/',
            },
            'features': [
                'JWT Token Authentication',
                'Blog Management with Image Upload',
                'SMS Template Management',
                'Customer Import/Export (Excel)',
                'Search and Pagination',
                'File Upload/Download',
                'Real-time Statistics',
            ]
        }
        
        return Response({
            'success': True,
            'data': documentation
        }, status=status.HTTP_200_OK)


# ======================================================================
# UTILITY API VIEWS
# ======================================================================

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def health_check(request):
    """
    Simple health check endpoint
    """
    return Response({
        'status': 'healthy',
        'timestamp': pd.Timestamp.now().isoformat(),
        'version': '1.0.0'
    }, status=status.HTTP_200_OK)